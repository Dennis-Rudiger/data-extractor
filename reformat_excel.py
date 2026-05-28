import pandas as pd
import re
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def format_excel(writer, df, sheet_name):
    worksheet = writer.sheets[sheet_name]
    
    # Freeze the top header row
    worksheet.freeze_panes = "A2"
    
    # Format Headers (Bold)
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    # Auto-adjust column widths and apply Number Formatting
    for idx, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        
        # Calculate max length
        max_len = max(
            df[col].astype(str).map(lambda x: len(str(x))).max(),
            len(str(col))
        ) + 2
        worksheet.column_dimensions[col_letter].width = max_len
        
        # Apply strict number formatting to float columns
        if pd.api.types.is_numeric_dtype(df[col]):
            for row in range(2, len(df) + 2):
                worksheet.cell(row=row, column=idx).number_format = '#,##0.00'

def process_sales(input_file, output_file):
    with open(input_file, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
    
    data = []
    headers = ["DATE", "SALES EXCL.", "SALES EXCEMPT", "V.A.T.", "TOTAL COSTS", "PROFIT", "G.P.%", "PETTY CASH", "CASH TAKEN", "LAYBYE PMNTS", "LAYBYE SALES", "EXPIRED REFUNDS"]
    
    for line in lines:
        parts = line.split()
        if len(parts) >= 12 and parts[0][0].isdigit() and len(parts[0]) == 6:
            data.append(parts[:12])
                
    if data:
        df = pd.DataFrame(data, columns=headers)
        
        # Fix date format
        df['DATE'] = pd.to_datetime(df['DATE'], format='%y%m%d', errors='coerce').dt.date
        
        # Convert numeric columns
        for col in df.columns:
            if col not in ["DATE"]:
                # Strip artifacts before conversion
                df[col] = df[col].astype(str).str.replace('******.**', '0.00', regex=False)
                df[col] = df[col].astype(str).str.replace(',', '', regex=False)
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Export with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sales Data')
            format_excel(writer, df, 'Sales Data')
            
        print(f"Organized and saved {output_file} with {len(df)} rows.")

def process_purchases(input_file, output_file):
    with open(input_file, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
    
    data = []
    current_supplier = ""
    
    for line in lines:
        if line.startswith('Supplier:'):
            # Extract just the supplier name
            parts = line.split('Period')
            current_supplier = parts[0].replace('Supplier:', '').strip()
            continue
            
        # Match lines that start with DD/MM/YYYY
        if re.match(r'^\d{2}/\d{2}/\d{4}', line):
            date_str = line[0:10].strip()
            ref = line[11:20].strip()
            details = line[21:55].strip()
            debit = line[55:69].strip()
            credit = line[69:85].strip()
            balance = line[85:].strip()
            
            data.append([current_supplier, date_str, ref, details, debit, credit, balance])

    if data:
        df = pd.DataFrame(data, columns=["Supplier", "Date", "Ref.", "Details", "Debit", "Credit", "Balance"])
        
        # Convert Date format
        df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce').dt.date
        
        # Convert to numeric where applicable
        for col in ["Debit", "Credit", "Balance"]:
            df[col] = df[col].astype(str).str.replace(',', '', regex=False)
            df[col] = pd.to_numeric(df[col], errors='coerce')
            df[col] = df[col].fillna(0) # optional: make NaN 0 for empty debits/credits? We'll leave as is.

        # Export with formatting
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Purchases Data')
            format_excel(writer, df, 'Purchases Data')
            
        print(f"Organized and saved {output_file} with {len(df)} rows.")

try:
    print("Reorganizing Sales and Purchases data...")
    process_sales('Sales 2023.txt', 'Sales 2023.xlsx')
    process_sales('Sales 2024.txt', 'Sales 2024.xlsx')
    process_purchases('Purchases 2023.txt', 'Purchases 2023.xlsx')
    process_purchases('Purchases 2024.txt', 'Purchases 2024.xlsx')
except Exception as e:
    print("Error:", e)
