import json
import pandas as pd
from openpyxl.styles import PatternFill, Font

# --- 1. Load Data ---
with open('bomas_average_moq.json', 'r', encoding='utf-8') as f:
    bomas = json.load(f)

with open('march_april_moq.json', 'r', encoding='utf-8') as f:
    march_april = json.load(f)

# --- 2. Build Basic MOQ Report ---
writer1 = pd.ExcelWriter('MOQ_Report_March_April.xlsx', engine='openpyxl')
for cat_name, cat_data in march_april['categories'].items():
    if not cat_data['products']: continue
    
    df = pd.DataFrame(cat_data['products'])
    df = df[['item_code', 'item_description', 'qty_out', 'daily_average', 'weekly_moq', 'buying_price', 'weekly_value']]
    df.columns = ['Item Code', 'Description', 'Qty Out', 'Daily Avg', 'Weekly MOQ', 'Buying Price', 'Weekly Value']
    df = df.sort_values('Qty Out', ascending=False)
    
    df.to_excel(writer1, sheet_name=cat_name[:31], index=False)
writer1.close()
print('Generated MOQ_Report_March_April.xlsx')


# --- 3. Build Comparison Report ---
# Map bomas products
b_map = {}
for cat_name, cat_data in bomas['categories'].items():
    for p in cat_data['products']:
        b_map[p['item_code']] = {
            'b_qty_out': p.get('qty_out', 0),
            'b_daily_avg': p.get('daily_average', 0),
            'b_moq': p.get('weekly_moq', 0)
        }

comp_data = []
for cat_name, cat_data in march_april['categories'].items():
    for m_p in cat_data['products']:
        code = m_p['item_code']
        b_p = b_map.get(code, {'b_qty_out': 0, 'b_daily_avg': 0, 'b_moq': 0})
        
        m_daily = m_p['daily_average']
        b_daily = b_p['b_daily_avg']
        
        variance = m_daily - b_daily
        status = 'Increased' if variance > 0.1 else ('Dropped' if variance < -0.1 else 'Unchanged')
        
        comp_data.append({
            'Category': cat_name,
            'Item Code': code,
            'Description': m_p['item_description'],
            'Prev Daily Avg (74 Days)': b_daily,
            'New Daily Avg (36 Days)': m_daily,
            'Variance (Daily)': variance,
            'Status': status,
            'Prev MOQ': b_p['b_moq'],
            'New MOQ': m_p['weekly_moq'],
            'Variance (MOQ)': m_p['weekly_moq'] - b_p['b_moq']
        })

comp_df = pd.DataFrame(comp_data)

writer2 = pd.ExcelWriter('MOQ_Comparison_Bomas.xlsx', engine='openpyxl')
for cat, group in comp_df.groupby('Category'):
    if group.empty: continue
    # Sort by absolute variance to see biggest changes first
    group['Abs_Var'] = group['Variance (Daily)'].abs()
    group = group.sort_values('Abs_Var', ascending=False).drop(columns=['Category', 'Abs_Var'])
    
    sheet_name = cat[:31]
    group.to_excel(writer2, sheet_name=sheet_name, index=False)
    
    # Simple formatting
    workbook = writer2.book
    worksheet = writer2.sheets[sheet_name]
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_col=10), start=2):
        stat_cell = row[5] # Status column (F) is index 5
        if stat_cell.value == 'Increased':
            stat_cell.fill = green_fill
        elif stat_cell.value == 'Dropped':
            stat_cell.fill = red_fill

writer2.close()
print('Generated MOQ_Comparison_Bomas.xlsx')

