import json
import os
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import inch
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Configuration
MONTHLY_TARGETS = {
    'OVERALL': 60_000_000,
    'PAINTS': 7_000_000,
    'PLUMBING': 3_000_000,
    'ELECTRICALS': 1_500_000,
    'GENERAL HARDWARE': 48_500_000,
}

VAT_RATE = 1.16

def load_data(filepath):
    with open(filepath, 'r') as f:
        return json.load(f)

# Helper to merge reps
def process_data(raw_data):
    reps = {}
    for cat, cat_reps in raw_data["categories"].items():
        for rep, data in cat_reps.items():
            rep_name = "Betha Odumo" if rep == "Magdalene" else rep
            if rep_name not in reps:
                reps[rep_name] = {'sales': 0, 'profit': 0, 'categories': {}}
            
            sales = data['sales_incl']
            profit = data['profit'] * VAT_RATE
            
            reps[rep_name]['sales'] += sales
            reps[rep_name]['profit'] += profit
            
            if cat not in reps[rep_name]['categories']:
                reps[rep_name]['categories'][cat] = {'sales': 0, 'profit': 0}
            reps[rep_name]['categories'][cat]['sales'] += sales
            reps[rep_name]['categories'][cat]['profit'] += profit
            
    return reps

def get_category_totals(raw_data):
    totals = {}
    for cat, cat_reps in raw_data["categories"].items():
        totals[cat] = sum(d['sales_incl'] for d in cat_reps.values())
    return totals

def main():
    print("Loading data...")
    w1_raw = load_data("sales_march_w1.json")
    w2_raw = load_data("sales_march_w2.json")
    
    w1_reps = process_data(w1_raw)
    w2_reps = process_data(w2_raw)
    
    w1_cat = get_category_totals(w1_raw)
    w2_cat = get_category_totals(w2_raw)
    
    # Combined Totals
    all_reps = set(w1_reps.keys()) | set(w2_reps.keys())
    all_cats = set(MONTHLY_TARGETS.keys()) - {'OVERALL'}
    
    combined_reps = {}
    for rep in all_reps:
        w1_s = w1_reps.get(rep, {}).get('sales', 0)
        w2_s = w2_reps.get(rep, {}).get('sales', 0)
        combined_reps[rep] = {
            'w1': w1_s,
            'w2': w2_s,
            'total': w1_s + w2_s,
            'growth': ((w2_s - w1_s) / w1_s * 100) if w1_s > 0 else 100
        }
        
    combined_cats = {}
    for cat in all_cats:
        w1_c = w1_cat.get(cat, 0)
        w2_c = w2_cat.get(cat, 0)
        combined_cats[cat] = {
            'w1': w1_c,
            'w2': w2_c,
            'total': w1_c + w2_c,
            'growth': ((w2_c - w1_c) / w1_c * 100) if w1_c > 0 else 100
        }
        
    w1_overall = sum(c['w1'] for c in combined_cats.values())
    w2_overall = sum(c['w2'] for c in combined_cats.values())
    total_overall = w1_overall + w2_overall
    
    os.makedirs("charts/march_midmonth", exist_ok=True)
    
    # Chart 1: Category Comparison
    labels = list(combined_cats.keys())
    w1_vals = [combined_cats[l]['w1'] for l in labels]
    w2_vals = [combined_cats[l]['w2'] for l in labels]
    
    x = np.arange(len(labels))
    width = 0.35
    
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(x - width/2, w1_vals, width, label='Week 1', color='#3498db')
    ax.bar(x + width/2, w2_vals, width, label='Week 2', color='#2ecc71')
    ax.set_ylabel('Sales (KES)')
    ax.set_title('Week 1 vs Week 2 Sales by Category')
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.legend()
    for i, v in enumerate(w1_vals):
        ax.text(i - width/2, v + 10000, f"{v/1e6:.1f}M", ha='center', va='bottom', fontsize=8)
    for i, v in enumerate(w2_vals):
        ax.text(i + width/2, v + 10000, f"{v/1e6:.1f}M", ha='center', va='bottom', fontsize=8)
    plt.tight_layout()
    plt.savefig("charts/march_midmonth/cat_comparison.png")
    plt.close()
    
    # Chart 2: Rep Comparison (Top 5)
    sorted_reps = sorted(combined_reps.items(), key=lambda x: x[1]['total'], reverse=True)[:6]
    rep_labels = [r[0] for r in sorted_reps]
    r_w1 = [r[1]['w1'] for r in sorted_reps]
    r_w2 = [r[1]['w2'] for r in sorted_reps]
    
    x_r = np.arange(len(rep_labels))
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.bar(x_r - width/2, r_w1, width, label='Week 1', color='#9b59b6')
    ax.bar(x_r + width/2, r_w2, width, label='Week 2', color='#e74c3c')
    ax.set_ylabel('Sales (KES)')
    ax.set_title('Week 1 vs Week 2 Sales by Top Reps')
    ax.set_xticks(x_r)
    ax.set_xticklabels(rep_labels, rotation=45, ha='right')
    ax.legend()
    plt.tight_layout()
    plt.savefig("charts/march_midmonth/rep_comparison.png")
    plt.close()

    # Chart 3: Progress towards Target
    cats = list(MONTHLY_TARGETS.keys())
    targets = [MONTHLY_TARGETS[c] for c in cats]
    achieved = [total_overall if c == 'OVERALL' else combined_cats.get(c, {}).get('total', 0) for c in cats]
    
    fig, ax = plt.subplots(figsize=(10, 6))
    y_pos = np.arange(len(cats))
    ax.barh(y_pos, targets, color='#ecf0f1', label='Monthly Target')
    ax.barh(y_pos, achieved, color='#34495e', label='Achieved (W1+W2)')
    ax.set_yticks(y_pos)
    ax.set_yticklabels(cats)
    ax.set_xlabel('Sales (KES)')
    ax.set_title('Cumulative Sales vs Monthly Target')
    for i, v in enumerate(achieved):
        pct = (v / targets[i]) * 100 if targets[i] > 0 else 0
        ax.text(v, i, f" {pct:.1f}%", va='center')
    ax.legend()
    plt.tight_layout()
    plt.savefig("charts/march_midmonth/target_progress.png")
    plt.close()

    print("Generating Reports...")
    # 1. PDF Report
    doc = SimpleDocTemplate("march_w1_w2_comparison.pdf", pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    title_style.alignment = TA_CENTER
    
    elements.append(Paragraph("March 2026: Mid-Month Sales Analysis (W1 vs W2)", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary Table
    elements.append(Paragraph(f"Overall Progress: KES {total_overall:,.2f} of KES {MONTHLY_TARGETS['OVERALL']:,.2f} ({total_overall/MONTHLY_TARGETS['OVERALL']*100:.1f}%)", styles['Heading2']))
    
    cat_data = [["Category", "Week 1", "Week 2", "Total (W1+W2)", "W1->W2 Growth", "Monthly Target", "% Achieved"]]
    for cat, data in combined_cats.items():
        target = MONTHLY_TARGETS.get(cat, 1)
        pct = (data['total'] / target) * 100
        growth_str = f"+{data['growth']:.1f}%" if data['growth'] > 0 else f"{data['growth']:.1f}%"
        cat_data.append([
            cat,
            f"KES {data['w1']:,.0f}",
            f"KES {data['w2']:,.0f}",
            f"KES {data['total']:,.0f}",
            growth_str,
            f"KES {target:,.0f}",
            f"{pct:.1f}%"
        ])
    
    t = Table(cat_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#ecf0f1')),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.5*inch))
    
    # Images
    elements.append(Image("charts/march_midmonth/cat_comparison.png", width=6*inch, height=4*inch))
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Image("charts/march_midmonth/target_progress.png", width=6*inch, height=4*inch))
    
    doc.build(elements)
    print("  -> march_w1_w2_comparison.pdf generated")

    # 2. Excel Report
    wb = Workbook()
    ws = wb.active
    ws.title = "W1 vs W2 Comparison"
    
    headers = ["Rep Name", "Week 1 Sales", "Week 2 Sales", "Total Sales", "Growth %"]
    ws.append(headers)
    
    # Format headers
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    for r_name, r_data in sorted(combined_reps.items(), key=lambda x: x[1]['total'], reverse=True):
        ws.append([
            r_name,
            r_data['w1'],
            r_data['w2'],
            r_data['total'],
            r_data['growth'] / 100.0  # Decimal for excel format
        ])
    
    for row in range(2, len(combined_reps) + 2):
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5).number_format = '0.00%'

    wb.save("march_w1_w2_comparison.xlsx")
    print("  -> march_w1_w2_comparison.xlsx generated")
    print("Done!")

if __name__ == '__main__':
    main()