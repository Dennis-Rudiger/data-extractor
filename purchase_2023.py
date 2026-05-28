import pandas as pd
import re
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def format_excel(writer, df, sheet_name):
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes = "A2"
    
    for cell in worksheet["1:1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    for idx, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        max_len = max(
            df[col].astype(str).map(len).max(),
            len(str(col))
        ) + 2
        worksheet.column_dimensions[col_letter].width = max_len
        
        if pd.api.types.is_numeric_dtype(df[col]):
            for row in range(2, len(df) + 2):
                worksheet.cell(row=row, column=idx).number_format = '#,##0.00'

def process_purchases(input_file, output_file):
    with open(input_file, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
    
    data = []
    current_supplier = ""  # ✅ Track supplier, attach it to EACH row — don't make it a row itself
    
    for line in lines:
        if line.startswith('Supplier:'):
            parts = line.split('Period')
            current_supplier = parts[0].replace('Supplier:', '').strip()
            continue  # ✅ Skip to next line, don't append anything yet
            
        if re.match(r'^\d{2}/\d{2}/\d{4}', line):
            data.append([
                current_supplier,       # ✅ Supplier on every data row
                line[0:10].strip(),     # Date
                line[11:20].strip(),    # Ref
                line[21:55].strip(),    # Details
                line[55:69].strip(),    # Debit
                line[69:85].strip(),    # Credit
                line[85:].strip()       # Balance
            ])

    if data:
        df = pd.DataFrame(data, columns=["Supplier", "Date", "Ref.", "Details", "Debit", "Credit", "Balance"])
        
        df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce').dt.date
        
        for col in ["Debit", "Credit", "Balance"]:
            df[col] = df[col].str.replace(',', '', regex=False)
            df[col] = pd.to_numeric(df[col], errors='coerce')

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Purchases Data')
            format_excel(writer, df, 'Purchases Data')
            
        print(f"Saved {output_file} with {len(df)} rows.")
    else:
        print("No data found.")

process_purchases('Purchases 2023.txt', 'Purchases fixed 2023.xlsx')